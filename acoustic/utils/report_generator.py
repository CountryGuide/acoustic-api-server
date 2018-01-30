import json
import string

from os import path
from datetime import datetime
from random import randint
from django.conf import settings

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side

from .noise_calculation import FREQUENCIES


def generate_report(options):
    data = json.loads(options)
    ac = Alignment(horizontal="center", vertical="center")
    ar = Alignment(horizontal="right", vertical="center", indent=1, wrap_text=True)
    border_lr = Border(left=Side(style='thin'),
                       right=Side(style='thin'))
    border_tlr = Border(top=Side(style='thin'),
                        left=Side(style='thin'),
                        right=Side(style='thin'))
    border_tr = Border(top=Side(style='thin'),
                       right=Side(style='thin'))
    border_br = Border(bottom=Side(style='thin'),
                       right=Side(style='thin'))
    border_blr = Border(bottom=Side(style='thin'),
                        left=Side(style='thin'),
                        right=Side(style='thin'))
    border_b = Border(bottom=Side(style='thin'))
    border_r = Border(right=Side(style='thin'))
    border_t = Border(top=Side(style='thin'))

    wb = Workbook()
    ws_rep = wb.active
    ws_rep.title = 'Report'

    ws_rep['A1'] = 'Report generated at:'
    ws_rep['A2'] = datetime.now()

    ws_rep['A4'] = 'Показатель'
    ws_rep['A6'] = 'Средний уровень ударного шума под перекрытием, дБ'
    ws_rep['A7'] = 'Время реверберации,с'
    ws_rep['A8'] = 'Приведенный уровень шума, дБ'
    ws_rep['A9'] = 'Оценочная кривая'
    ws_rep['A10'] = 'Неблагоприятные отклонения'
    ws_rep['A11'] = 'Сумма неблагоприятных отклонений'
    ws_rep['A12'] = 'Целое число децибел, добавл./выч. из норм.кривой:'
    ws_rep['A14'] = 'Смещеная оценочная кривая, дБ'
    ws_rep['A15'] = 'Неблагоприятные отклонения'
    ws_rep['A16'] = 'Сумма неблагоприятных отклонений с учетом смещения оценочной кривой'
    ws_rep['A17'] = 'Индекс приведенного ударного шума'

    ws_rep['B4'] = 'Среднегеометрические частоты третьоктавных полос, Гц'

    ws_rep.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)
    ws_rep.merge_cells(start_row=4, start_column=2, end_row=4, end_column=17)

    for i in range(16):
        ws_rep.cell(row=5, column=(2 + i), value=FREQUENCIES[i])
        ws_rep.cell(row=6, column=(2 + i), value=data['average'][i])
        ws_rep.cell(row=7, column=(2 + i), value=data['reverberation-time'][i])
        ws_rep.cell(row=8, column=(2 + i), value=data['reduced'][i])
        ws_rep.cell(row=9, column=(2 + i), value=data['evaluation_curve'][i])
        ws_rep.cell(row=10, column=(2 + i), value=data['initial_deltas'][i])
        ws_rep.cell(row=14, column=(2 + i), value=data['reduced_evaluation_curve'][i])
        ws_rep.cell(row=15, column=(2 + i), value=data['deltas'][i])

    ws_rep['B11'] = data['initial_deltas_sum']
    ws_rep['B12'] = data['dB_difference']
    ws_rep['B16'] = data['deltas_sum']
    ws_rep['B17'] = data['reduced_noise_index']

    ws_rep['A4'].alignment = ac
    ws_rep['A1'].alignment = ac
    ws_rep['A2'].alignment = ac

    for i in range(15):
        for j in range(16):
            ws_rep.cell(row=(3 + i), column=(2 + j)).alignment = ac

    for i in range(13):
        ws_rep.cell(row=(5 + i), column=1).alignment = ar

    ws_rep['A4'].border = border_tlr
    ws_rep['A5'].border = border_blr
    ws_rep['A17'].border = border_blr
    ws_rep['Q4'].border = border_tr
    ws_rep['Q5'].border = border_br
    ws_rep['Q17'].border = border_br

    for i in range(11):
        ws_rep.cell(row=(6 + i), column=1).border = border_lr
        ws_rep.cell(row=(6 + i), column=17).border = border_r

    for i in range(15):
        ws_rep.cell(row=4, column=(i + 2)).border = border_t
        ws_rep.cell(row=5, column=(i + 2)).border = border_b
        ws_rep.cell(row=17, column=(i + 2)).border = border_b

    ws_rep.column_dimensions['A'].width = 56.43

    for i in string.ascii_uppercase[1:17]:
        ws_rep.column_dimensions[i].width = 8

    now = str(datetime.date(datetime.now()))
    salt = str(randint(1000, 9999))
    file_name = 'Report_' + now + '_' + salt + '.xlsx'
    file_path = path.join(settings.MEDIA_ROOT, file_name)

    wb.save(file_path)

    return json.dumps({'path': file_name})
